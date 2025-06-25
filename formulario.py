import streamlit as st
from datetime import date
from openpyxl import load_workbook
import pandas as pd
import os

# --- Inicializações de sessão ---
if "insumos" not in st.session_state:
    st.session_state.insumos = []
if "resetar_insumo" not in st.session_state:
    st.session_state.resetar_insumo = False
if "resetar_pedido" not in st.session_state:
    st.session_state.resetar_pedido = False

# --- Funções auxiliares ---
def resetar_campos_insumo():
    st.session_state.resetar_insumo = True

def resetar_formulario():
    st.session_state.resetar_pedido = True
    resetar_campos_insumo()
    st.session_state.insumos = []

def registrar_historico(numero, obra, data):
    historico_path = "historico_pedidos.csv"
    registro = {"numero": str(numero).strip(), "obra": str(obra).strip(), "data": data.strftime("%Y-%m-%d")}
    if os.path.exists(historico_path):
        df_hist = pd.read_csv(historico_path, dtype=str)
        if not ((df_hist["numero"] == registro["numero"]) & (df_hist["obra"] == registro["obra"])).any():
            df_hist = pd.concat([df_hist, pd.DataFrame([registro])], ignore_index=True)
            df_hist.to_csv(historico_path, index=False, encoding="utf-8")
            st.success(f"📌 Pedido {numero} adicionado ao histórico.")
        else:
            st.warning(f"ℹ️ Pedido {numero} já está no histórico.")
    else:
        df_hist = pd.DataFrame([registro])
        df_hist.to_csv(historico_path, index=False, encoding="utf-8")
        st.success(f"📌 Histórico criado com o pedido {numero}.")

def carregar_dados():
    df_empreend = pd.read_excel("Empreendimentos.xlsx")
    df_insumos = pd.read_excel("Insumos.xlsx")
    df_empreend.loc[-1] = ["", "", "", ""]
    df_empreend.index = df_empreend.index + 1
    df_empreend = df_empreend.sort_index()
    insumos_vazios = pd.DataFrame({"Código": [""], "Descrição": [""], "Unidade": [""]})
    df_insumos = pd.concat([insumos_vazios, df_insumos], ignore_index=True)
    return df_empreend, df_insumos

# --- Carrega dados ---
df_empreend, df_insumos = carregar_dados()

# --- Logo e título ---
st.image("logo.png", width=300)
st.markdown("""
    <div style='text-align: center;'>
        <h2 style='color: #003366;'>Sistema de Pedidos de Materiais</h2>
        <p style='font-size: 14px; color: #555;'>Preencha os campos com atenção. Evite abreviações desnecessárias.<br>
        </p>
    </div>
""", unsafe_allow_html=True)

# --- Dados do Pedido ---
with st.expander("📋 Dados do Pedido", expanded=True):
    if st.session_state.resetar_pedido:
        st.session_state.pedido_numero = ""
        st.session_state.data_pedido = date.today()
        st.session_state.solicitante = ""
        st.session_state.executivo = ""
        st.session_state.obra_selecionada = ""
        st.session_state.cnpj = ""
        st.session_state.endereco = ""
        st.session_state.cep = ""
        st.session_state.resetar_pedido = False

    col1, col2 = st.columns(2)
    with col1:
        pedido_numero = st.text_input("Pedido Nº", key="pedido_numero")
        solicitante = st.text_input("Solicitante", key="solicitante")
        obra_selecionada = st.selectbox("Obra", df_empreend["NOME"].unique(), index=0, key="obra_selecionada")
    with col2:
        data_pedido = st.date_input("Data", value=st.session_state.get("data_pedido", date.today()), key="data_pedido")
        executivo = st.text_input("Executivo", key="executivo")

    if obra_selecionada:
        dados_obra = df_empreend[df_empreend["NOME"] == obra_selecionada].iloc[0]
        st.session_state.cnpj = dados_obra["EMPRD_CNPJFAT"]
        st.session_state.endereco = dados_obra["ENDEREÇO"]
        st.session_state.cep = dados_obra["Cep"]

    st.text_input("CNPJ/CPF", value=st.session_state.get("cnpj", ""), disabled=True)
    st.text_input("Endereço", value=st.session_state.get("endereco", ""), disabled=True)
    st.text_input("CEP", value=st.session_state.get("cep", ""), disabled=True)

st.divider()

# --- Adição de Insumos ---
with st.expander("➕ Adicionar Insumo", expanded=True):
    if st.session_state.resetar_insumo:
        st.session_state.descricao = ""
        st.session_state.descricao_livre = ""
        st.session_state.codigo = ""
        st.session_state.unidade = ""
        st.session_state.quantidade = 0.0
        st.session_state.complemento = ""
        st.session_state.resetar_insumo = False

    descricao = st.selectbox("Descrição do insumo (Digite em Maiúsculo)", df_insumos["Descrição"].unique(), key="descricao")
    codigo = ""
    unidade = ""
    if descricao:
        dados_insumo = df_insumos[df_insumos["Descrição"] == descricao].iloc[0]
        codigo = dados_insumo["Código"]
        unidade = dados_insumo["Unidade"]
    
    st.write("Ou preencha manualmente se não estiver listado:")
    descricao_livre = st.text_input("Nome do insumo (livre)", key="descricao_livre")
    
    # Verifica se insumo foi selecionado da base
    usando_base = bool(descricao and not descricao_livre)
    
    # Define código e unidade com base no tipo de entrada
    if usando_base:
        dados_insumo = df_insumos[df_insumos["Descrição"] == descricao].iloc[0]
        codigo = dados_insumo["Código"]
        unidade = dados_insumo["Unidade"]
    else:
        codigo = ""
        unidade = ""
    
    # Campo código sempre bloqueado
    st.text_input("Código do insumo", value=codigo, key="codigo", disabled=True)
    
    # Campo unidade apenas editável se for insumo manual
    unidade = st.text_input("Unidade", value=unidade, key="unidade", disabled=usando_base)

    quantidade = st.number_input("Quantidade", min_value=0.0, format="%.2f", key="quantidade")
    complemento = st.text_area("Complemento", key="complemento")

    if st.button("➕ Adicionar insumo"):
        descricao_final = descricao if descricao else descricao_livre
        usando_base = bool(descricao)  # True se o insumo veio do selectbox
    
        if descricao_final and quantidade > 0 and (usando_base or unidade.strip()):
            novo_insumo = {
                "descricao": descricao_final,
                "codigo": codigo if usando_base else "",
                "unidade": unidade,
                "quantidade": quantidade,
                "complemento": complemento,
            }
            st.session_state.insumos.append(novo_insumo)
            st.success("Insumo adicionado com sucesso!")
            resetar_campos_insumo()
            st.rerun()
        else:
            st.warning("Preencha todos os campos obrigatórios do insumo.")

# --- Renderiza tabela de insumos ---
if st.session_state.insumos:
    st.subheader("📦 Insumos adicionados")
    for i, insumo in enumerate(st.session_state.insumos):
        cols = st.columns([6, 1])
        with cols[0]:
            st.markdown(f"**{i+1}.** {insumo['descricao']} — {insumo['quantidade']} {insumo['unidade']}")
        with cols[1]:
            if st.button("🗑️", key=f"delete_{i}"):
                st.session_state.insumos.pop(i)
                st.rerun()

# --- Finalização do Pedido ---
if st.button("📤 Enviar Pedido"):
    campos_obrigatorios = [
        st.session_state.pedido_numero,
        st.session_state.data_pedido,
        st.session_state.solicitante,
        st.session_state.executivo,
        st.session_state.obra_selecionada,
        st.session_state.cnpj,
        st.session_state.endereco,
        st.session_state.cep
    ]

    if not all(campos_obrigatorios):
        st.warning("⚠️ Preencha todos os campos obrigatórios antes de enviar o pedido.")
        st.stop()

    try:
        caminho_modelo = "Modelo_Pedido.xlsx"
        wb = load_workbook(caminho_modelo)
        ws = wb["Pedido"]

        ws["H2"] = st.session_state.pedido_numero
        ws["D3"] = st.session_state.data_pedido.strftime("%d/%m/%Y")
        ws["D4"] = st.session_state.solicitante
        ws["D5"] = st.session_state.executivo
        ws["D7"] = st.session_state.obra_selecionada
        ws["D8"] = st.session_state.cnpj
        ws["D9"] = st.session_state.endereco
        ws["D10"] = st.session_state.cep

        linha = 13
        for insumo in st.session_state.insumos:
            ws[f"C{linha}"] = insumo["codigo"]
            ws[f"D{linha}"] = insumo["descricao"]
            ws[f"F{linha}"] = insumo["unidade"]
            ws[f"G{linha}"] = insumo["quantidade"]
            ws[f"H{linha}"] = insumo["complemento"]
            linha += 1

        nome_saida = f"pedido_{st.session_state.pedido_numero or 'sem_numero'}_{st.session_state.obra_selecionada}.xlsx"
        wb.save(nome_saida)

        with open(nome_saida, "rb") as f:
            excel_bytes = f.read()

        st.success("✅ Pedido gerado com sucesso!")
        st.download_button("📥 Baixar Excel", data=excel_bytes, file_name=nome_saida)

        st.info("Para gerar um PDF, abra o arquivo no Excel e use a opção 'Salvar como PDF'.")

        numero = st.session_state.pedido_numero
        obra = st.session_state.obra_selecionada
        data_pedido = st.session_state.data_pedido
        
        registrar_historico(numero, obra, data_pedido)
        resetar_formulario()

    except Exception as e:
        st.error(f"Erro ao gerar pedido: {e}")

# --- Histórico de pedidos ---
if st.checkbox("📖 Ver histórico de pedidos"):
    historico_path = "historico_pedidos.csv"
    if os.path.exists(historico_path):
        df = pd.read_csv(historico_path, dtype={"numero": str, "obra": str, "data": str})
    
        if "data" in df.columns:
            try:
                df["data"] = pd.to_datetime(df["data"], errors="coerce")
                df = df[df["data"].notna()]
            except Exception as e:
                st.error(f"Erro ao processar datas: {e}")
                st.stop()
        else:
            st.error("A coluna 'data' não foi encontrada no histórico.")
            st.stop()

        obra_filtro = st.selectbox("Filtrar por obra", ["Todas"] + sorted(df["obra"].dropna().unique()))
        mes_filtro = st.selectbox("Filtrar por mês", ["Todos"] + sorted(df["data"].dt.strftime("%Y-%m").unique()))

        if obra_filtro != "Todas":
            df = df[df["obra"] == obra_filtro]
        if mes_filtro != "Todos":
            df = df[df["data"].dt.strftime("%Y-%m") == mes_filtro]

        st.table(df[["numero", "obra", "data"]])
    else:
        st.info("Nenhum pedido registrado ainda.")
